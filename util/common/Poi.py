import openpyxl
from win32com.client import Dispatch
from openpyxl import Workbook


class Poi:
    # 表
    sheet = None

    workBook = None
    # 表名
    file_name = None

    def __init__(self, file_name: str, sheet_name=None, data_only=True):
        self.file_name = file_name
        if sheet_name is None:
            self.workBook = Workbook()
            self.sheet = self.workBook.active
        else:
            self.workBook = openpyxl.load_workbook(file_name, data_only=data_only)
            self.sheet = self.workBook[sheet_name]

    def get_sheet(self):
        return self.sheet

    def max_row(self):
        return self.sheet.max_row

    def max_column(self):
        return self.sheet.max_column

    def read_value(self, row_num: int, column_num: int):
        value = self.sheet.cell(column=column_num, row=row_num).value
        if value is None:
            return ""
        else:
            return value

    def write_value(self, row_num: int, column_num: int, value: str):
        self.sheet.cell(column=column_num, row=row_num, value=value)

    def save_file(self):
        self.workBook.save(self.file_name)

    def get_cell(self, row_num: int, column_num: int):
        return self.sheet.cell(column=column_num, row=row_num)

    def foreach(self, foreach, start_row=1, start_column=1):
        for row_num in range(start_row, self.max_row() + 1):
            for column_num in range(start_column, self.max_column() + 1):
                foreach(row_num, column_num, self.get_cell(column_num=column_num, row_num=row_num))
            print()

    def foreach_row(self, foreach, start_row=1, end_row=None):
        if end_row is None:
            end_row = self.max_row()
        for row_num in range(start_row, end_row + 1):
            data = []
            for column_num in range(1, self.max_column() + 1):
                data.append(self.read_value(column_num=column_num, row_num=row_num))
            foreach(data)

    def just_open(self, filename):
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(filename)
        xlBook.Save()
        xlBook.Close()
